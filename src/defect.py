from openpyxl import load_workbook
import api
import os

def main():
    list = []

    for a in range(1,7):
        dev = '{}{}'.format("dev", a)
        root = api.path + '%s/defects/' % dev
        dirlist = [item for item in os.listdir(root) if os.path.isfile(os.path.join(root, item))]
        for i in dirlist:
            wb = load_workbook(root + i)
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            for row in range(2,worksheet.max_row):
                for column in "ABCDEFGHJ":
                    cell_name = "{}{}".format(column, row)
                    list.append(worksheet[cell_name].value)
                if not list[0]:
                    list = []
                    break
                if not list[2]:
                    list[2] = ""
                if list[2] == '':
                    list[2] = 'NONE'
                data_defect = {
                    "created": str(list[5])[:-9],
                    "description": list[8],
                    "developer": api.get('developers', a),
                    "id": "",
                    "modified": str(list[6])[:-9],
                    "priority": list[4].upper(),
                    "resolution": list[2].upper(),
                    "severity": list[3].upper(),
                    "summary": list[7],
                    "ticket": list[0],
                    "status": "PRIVATE",
                    "author": "",
                    "license": ""
                }
                api.request("Defect", 'defects', data_defect)
                list = []